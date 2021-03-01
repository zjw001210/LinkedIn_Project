# 要做的事情就是格式化为一条条的job数据，供我们判断
import pandas as pd
import openpyxl

def Title():
    a='1'
    ws['A'+a] = 'id'
    ws['B'+a] = '生产/成本'
    ws['C'+a] = '稳定'
    ws['D'+a] = '舒适'
    ws['E'+a] = '驾驶'
    ws['F'+a] = '安全'
    ws['G'+a] = '未知技术'
    ws['H'+a] = '非技术'
    ws['I'+a] = '支持'
    ws['J'+a] = '未知'
    ws['K'+a] = '存疑'
    ws['L'+a] = 'name'
    ws['M' + a] = 'job'
    ws['N'+a] = 'description'
    

filename=['0_15735', '15735_31500', '31500_47250', '47250_63000', '63000_79000', 
    '79000_95000','95000_110750', '110750_122500', '122500_138235', '138235_154000',
    '154000_170000', '170000_185735', '185735_','overdata']

for s in range(len(filename)):
    filepath_all='data_analysis_all/{}_da_all.xlsx'.format(filename[s])
    
    print('s=',filename[s])
    filepath_da_2='data_analysis_classify/{}_classify.xlsx'.format(filename[s])
    wb = openpyxl.load_workbook(filename=filepath_da_2)
    openpyxl.sheet_ranges = wb['Sheet1']
    ws = wb.active
    Title() # 运行Title()函数
    frame0 = pd.read_excel(filepath_all)

    c=0
    idx = 0
    for i in frame0.index:
        f = dict(frame0.loc[i])
            # print(f)
            # 把表格的每一行用index进行读取，转化成‘字典’便于操作

        idn1 = str(idx+2)
        idn2 = str(int(idn1)+1)
        ws['A'+idn1] = str(f['id'])
        ws['L'+idn1] = str(f['name'])
        ws['M' + idn1] = str(f['last_job'])
        ws['N'+idn1] = str(f['last_dsp'])
        ws['A'+idn2] = str(f['id'])
        ws['L'+idn2] = str(f['name'])
        ws['M' + idn2] = str(f['next_job'])
        ws['N'+idn2] = str(f['next_dsp'])
        idx += 2
    wb.save(filepath_da_2)