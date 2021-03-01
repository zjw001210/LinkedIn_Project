import pandas as pd
import openpyxl
# 调用python的类openpyxl，便于每个数据的填写
import re

def Title():
    a='1'
    ws['A'+a] = 'index'
    ws['B'+a] = 'gap'
    ws['C' + a] = 'jhlast_time'
    ws['D'+a] = 'jhnext_time'
    ws['E'+a] = 'id'
    ws['F'+a] = 'name'
    ws['G'+a] = 'last_company'
    ws['H'+a] = 'next_company'
    ws['I'+a] = 'last_job'
    ws['J'+a] = 'next_job'
    ws['K'+a] = 'last_dsp'
    ws['L'+a] = 'next_dsp'
    ws['M'+a] = 'boolean_result'
'''
filename=['0_15735', '15735_31500', '31500_47250', '47250_63000', '63000_79000', 
    '79000_95000','95000_110750', '110750_122500', '122500_138235', '138235_154000',
    '154000_170000', '170000_185735', '185735_']
'''
filename=['overdata']
for s in range(len(filename)):
    filepath_all='select_result_test/{}_index.xlsx'.format(filename[s])
    for gap in range(24):  #loop:生成每个数据表，按照gap=0到num-2
        print('s=',filename[s],' gap=',gap)
        filepath_jobhopping='jobhopping_result/{}/{}_jh_gap{}.xlsx'.format(filename[s],filename[s],gap)
        wb = openpyxl.load_workbook(filename=filepath_jobhopping)
        openpyxl.sheet_ranges = wb['Sheet1']
        ws = wb.active
        Title() # 运行Title()函数
        frame0 = pd.read_excel(filepath_all)
    
        c=0
        max_num=25 #最大的数字，是overdata的时候需要改成25，否则为12
        up = gap + 1
        idx = 0
        for i in frame0.index:
            f = dict(frame0.loc[i])
            # print(f)
            # 把表格的每一行用index进行读取，转化成‘字典’便于操作
            j = 1
            while j<=(max_num-up):
                clast_x = 'company_' + str(j+up)
                cnext_x = 'company_' + str(j)
                jlast_x = 'job_' + str(j+up)
                jnext_x = 'job_' + str(j)
                dlast_x = 'duration_' + str(j+up)
                dnext_x = 'duration_' + str(j)
                deslast_x = 'description_' + str(j+up)
                desnext_x = 'description_' + str(j)
                # 创造一系列代表title的变量值 
                # 顺序变量，从1开始，根据idx值变化，便于竖向索引
                dur_lastpart = re.split("-",str(f[dlast_x]))
                dur_nextpart = re.split("-",str(f[dnext_x]))
                if len(dur_lastpart)==1:
                    dur_lastpart.append('nan')
                if len(dur_nextpart)==1:
                    dur_nextpart.append('nan')
                #print('last',dur_lastpart)
                #print('next',dur_nextpart)
                idn = str(idx+2)
                ws['A'+idn] = idx
                ws['B'+idn] = gap
                ws['C' + idn] = dur_lastpart[1]
                ws['D'+idn] = dur_nextpart[0]
                ws['E'+idn] = 'overdata'+'_'+str(f['Unnamed: 0']) #个人id
                ws['F'+idn] = str(f['full_name'])
                ws['G'+idn] = str(f[clast_x])
                ws['H'+idn] = str(f[cnext_x])
                ws['I'+idn] = str(f[jlast_x])
                ws['J'+idn] = str(f[jnext_x])
                ws['K'+idn] = str(f[deslast_x])
                ws['L'+idn] = str(f[desnext_x])
                ws['M'+idn] = str(f['boolean_result'])
                idx += 1
                j += 1
        wb.save(filepath_jobhopping)
